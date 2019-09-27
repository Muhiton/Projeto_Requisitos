import sys, csv, xlsxwriter, pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook


class csv_para_excel: #Classe para transformar o CSV em XLSX.

    def __init__ (self):#Funcao inicial da classe.
        return #Continuar sem necessidade de chamada.

    def tratar_csv(csv_sem_tratar): #Funcao que trata o CSV com ";" e troca para ",".
        with open(csv_sem_tratar, "rt") as fin: #Abre o arquivo CSV sem tratamento.
            with open(r"C:\Users\Administrador\Desktop\Projeto_Requisitos\CSV\CSV_Out\CSVTratado.csv", "wt") as fout: #Cria o arquivo CSV tratado.
                for line in fin: #Repetição para percorrer o texto.
                    fout.write(line.replace(";", ",")) #Altera o ";" por ",".
        global csv_file #Cria a variavel global csv_file.
        csv_file = (r"C:\Users\Administrador\Desktop\Projeto_Requisitos\CSV\CSV_Out\CSVTratado.csv") #Armazena na variavel global o PATH do CSV tratado.
        return csv_file #Retorna a variavel para o controlador.

    def getdf(csv_file):#Extrai os dados do CSV tratado e armazena em um objeto Pandas.
        global data_frame #Cria a variavel global data_frame
        data_frame = pd.read_csv(csv_file, encoding="ISO-8859-1", header=0, index_col=0 ) # Com a funcao read_csv do Pandas abre o arquivo , que ja extrai os dados CSV e transforma em um Objeto.
        return data_frame #Retorna o objeto para o controlador.

    def postxlsx(data_frame): #Função que pega o objeto (data_frame) com o Data Frame (DF) do Pandas.
        print(data_frame) #Printa o DF para checagem visual.
        with pd.ExcelWriter(r"C:\Users\Administrador\Desktop\Projeto_Requisitos\XLSX\xlsxparacarga.xlsx", engine="openpyxl", mode="a") as writer: # Com a metodo ExcelWriter abre o arquivo xlsx aonde se deseja inserir os dados, com a biblioteca openpyxl, no modo (a) de "append", como writer. Notas:Provavelmente o "a" é disso, pesquisar.
            book = load_workbook(r"C:\Users\Administrador\Desktop\Projeto_Requisitos\XLSX\xlsxparacarga.xlsx")
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            data_frame.to_excel(writer, sheet_name='Pasta1', startrow=0, startcol=0) #Com o metodo to_excel ele passa o objeto (DF) do pandas, para o local especificado e com os parametros da engine e modo.
            print("Planilha pronta.")#Printa o exito da insercao.

    """
    import pandas as pd
    from openpyxl import load_workbook

    book = load_workbook('test.xlsx')
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name='tab_name', other_params)

    writer.save()
    """




    if __name__ == '__main__': #Controlador das funões da classe.
        tratar_csv(r"C:\Users\Administrador\Desktop\Projeto_Requisitos\CSV\CSV_In\CSVTesteSemTratamento.csv")#Chama a funcao que trata o CSV e como parametro a PAth do CSV sem tratamento.
        getdf(csv_file) #Chama a funcap que gera o data_frame, objeto com os dados CSV do pandas. Passando como parametro a PATH do CSV tratado.
        postxlsx(data_frame)#Chama a funcao


csv_para_excel() #Chama a classe que da carga dos dados CSV no EXCEL
